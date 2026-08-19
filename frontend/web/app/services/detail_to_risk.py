"""Convert a filled 상세결과보고서.xlsx into a 위험분석 및 평가 보고서 workbook."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

from app.config import EXPORT_DIR
from app.services.excel_export import (
    ExportResult,
    _export_os_group_key,
    generate_risk_workbook,
)

_CODE_RE = re.compile(r"^[A-Za-z]+-\d+$")
_SUMMARY_CODE_REF = re.compile(
    r"요약\s*통계['\"]?\s*!\s*\$?[A-Z]+\$?(\d+)",
    re.IGNORECASE,
)
_RESERVED_SHEETS = {
    "표지",
    "개정이력",
    "개요",
    "점검대상",
    "요약 통계",
    "보안수준 통계",
    "기준 DB",
    "sample",
    "매크로",
    "Origin(숨김처리)",
    "위험관리계획기준",
    "위험분석및평가기준",
    "잠재위험 및 대응책 DB",
}


@dataclass
class ParsedDiag:
    code: str
    status: str
    evidence: str = ""
    inspection_status: str = ""
    title: str = ""
    category: str = ""
    severity: str = ""
    processed_config: str = ""
    raw_config: str = ""
    vulnerable_config: str = ""
    err_msg: str = ""


@dataclass
class ParsedHost:
    hostname: str
    ip_address: str = ""
    target_os: str = "Linux"
    usage: str = ""
    cia_c: int = 3
    cia_i: int = 3
    cia_a: int = 3
    diagnostics: list[ParsedDiag] = field(default_factory=list)


def list_detailed_exports(limit: int = 25) -> list[str]:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    files = [
        p
        for p in EXPORT_DIR.glob("*.xlsx")
        if "상세결과" in p.name and not p.name.startswith("~$")
    ]
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return [p.name for p in files[:limit]]


def _nfc(value: str) -> str:
    return unicodedata.normalize("NFC", value or "")


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, CellRichText):
        text = str(value).strip()
    else:
        text = str(value).strip()
    if text.startswith("="):
        return ""
    return text


def _as_int(value, default: int = 3) -> int:
    try:
        n = int(value)
        if 1 <= n <= 3:
            return n
    except (TypeError, ValueError):
        pass
    return default


def _sheet_by_name(wb, name: str):
    if not name:
        return None
    if name in wb.sheetnames:
        return wb[name]
    target = _nfc(name)
    for sheet_name in wb.sheetnames:
        if _nfc(sheet_name) == target:
            return wb[sheet_name]
    truncated = target[:31]
    if truncated in wb.sheetnames:
        return wb[truncated]
    for sheet_name in wb.sheetnames:
        if sheet_name in _RESERVED_SHEETS:
            continue
        sn = _nfc(sheet_name)
        if target.startswith(sn) or sn.startswith(target):
            return wb[sheet_name]
    return None


def _result_to_status(result: str) -> str:
    text = (result or "").strip()
    if text in {"양호", "Y", "Pass", "PASS", "pass"}:
        return "Pass"
    if text in {"취약", "N", "Fail", "FAIL", "fail"}:
        return "Fail"
    return "N/A"


def _load_summary_codes(wb) -> dict[int, str]:
    mapping: dict[int, str] = {}
    if "요약 통계" not in wb.sheetnames:
        return mapping
    ws = wb["요약 통계"]
    for row in range(6, min((ws.max_row or 6) + 1, 80)):
        for col in (2, 3):
            code = _cell_text(ws.cell(row, col).value)
            if _CODE_RE.match(code):
                mapping[row] = code.upper()
                break
    return mapping


def _item_code(ws, row: int, summary_codes: dict[int, str]) -> str:
    cell = ws.cell(row, 3)
    text = _cell_text(cell.value)
    if _CODE_RE.match(text):
        return text.upper()
    raw = cell.value
    if isinstance(raw, str) and raw.startswith("="):
        compact = raw.replace(" ", "")
        m = _SUMMARY_CODE_REF.search(compact)
        if m:
            return summary_codes.get(int(m.group(1)), "")
        m = re.search(r"\$B(\d+)", raw, re.I)
        if m:
            return summary_codes.get(int(m.group(1)), "")
    return ""


def _host_identity(ws) -> tuple[str, str, str]:
    hostname = _cell_text(ws["D4"].value) or _cell_text(ws["C6"].value) or ws.title
    ip_address = _cell_text(ws["D5"].value) or _cell_text(ws["C7"].value)
    target_os = _cell_text(ws["H4"].value) or _cell_text(ws["I6"].value) or "Linux"
    return hostname, ip_address, target_os


def _load_criteria(wb) -> dict[str, dict[str, str]]:
    mapping: dict[str, dict[str, str]] = {}
    if "기준 DB" not in wb.sheetnames:
        return mapping
    ws = wb["기준 DB"]
    for row in range(2, (ws.max_row or 2) + 1):
        code = _cell_text(ws.cell(row, 1).value).upper()
        if not code:
            continue
        mapping[code] = {
            "title": _cell_text(ws.cell(row, 2).value),
            "pass": _cell_text(ws.cell(row, 3).value),
            "fail": _cell_text(ws.cell(row, 4).value),
            "na": _cell_text(ws.cell(row, 5).value),
        }
    return mapping


def _compose_status(
    result: str,
    code: str,
    criteria: dict[str, dict[str, str]],
    pass_comment: str,
    fail_comment: str,
) -> str:
    rec = criteria.get(code.upper()) or {}
    text = (result or "").strip()
    if text in {"양호", "Y"}:
        return (rec.get("pass") or "") + (pass_comment or "")
    if text in {"취약", "N"}:
        return (rec.get("fail") or "") + ((fail_comment or pass_comment) or "")
    return rec.get("na") or ""


def _iter_target_rows(ws):
    for row in range(7, (ws.max_row or 7) + 1):
        hostname = _cell_text(ws.cell(row, 2).value)
        if not hostname or hostname.lower() == "sample":
            continue
        yield {
            "hostname": hostname,
            "target_os": _cell_text(ws.cell(row, 3).value) or "Linux",
            "ip_address": _cell_text(ws.cell(row, 4).value),
            "usage": _cell_text(ws.cell(row, 5).value),
            "cia_c": _as_int(ws.cell(row, 6).value),
            "cia_i": _as_int(ws.cell(row, 7).value),
            "cia_a": _as_int(ws.cell(row, 8).value),
        }


def _host_sheet_names(wb) -> list[str]:
    names = []
    for name in wb.sheetnames:
        if name in _RESERVED_SHEETS:
            continue
        if name.lower() == "sample":
            continue
        names.append(name)
    return names


def _parse_host_items(
    ws,
    criteria: dict[str, dict[str, str]],
    cached_status: dict[tuple[str, int], str] | None = None,
    summary_codes: dict[int, str] | None = None,
) -> list[ParsedDiag]:
    items: list[ParsedDiag] = []
    seen: set[str] = set()
    cached_status = cached_status or {}
    summary_codes = summary_codes or {}
    max_row = min(ws.max_row or 28, 140)
    for row in range(20, max_row + 1):
        code = _item_code(ws, row, summary_codes)
        if not _CODE_RE.match(code):
            continue
        key = code.upper()
        if key in seen:
            continue
        seen.add(key)
        result = _cell_text(ws.cell(row, 15).value)
        pass_comment = _cell_text(ws.cell(row, 21).value)
        fail_comment = _cell_text(ws.cell(row, 22).value)
        inspection_status = cached_status.get((ws.title, row), "")
        if not inspection_status:
            inspection_status = _cell_text(ws.cell(row, 16).value)
        if not inspection_status:
            inspection_status = _compose_status(
                result, key, criteria, pass_comment, fail_comment
            )
        items.append(
            ParsedDiag(
                code=key,
                status=_result_to_status(result),
                evidence=inspection_status,
                inspection_status=inspection_status,
                title=_cell_text(ws.cell(row, 4).value),
            )
        )
    return items


def _load_cached_inspection_status(path: Path) -> dict[tuple[str, int], str]:
    cached: dict[tuple[str, int], str] = {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            for name in wb.sheetnames:
                ws = wb[name]
                for row in ws.iter_rows(min_row=20, max_row=140, min_col=16, max_col=16):
                    cell = row[0]
                    val = cell.value
                    if val is None:
                        continue
                    text = str(val).strip()
                    if text and not text.startswith("="):
                        cached[(name, cell.row)] = text
        finally:
            wb.close()
    except Exception:
        return {}
    return cached


def parse_detailed_report(path: Path) -> list[ParsedHost]:
    wb = load_workbook(path, data_only=False, rich_text=True)
    try:
        criteria = _load_criteria(wb)
        summary_codes = _load_summary_codes(wb)
        cached_status = _load_cached_inspection_status(path)
        hosts: list[ParsedHost] = []
        used_sheets: set[str] = set()

        if "점검대상" in wb.sheetnames:
            for row in _iter_target_rows(wb["점검대상"]):
                sheet = _sheet_by_name(wb, row["hostname"])
                items = (
                    _parse_host_items(sheet, criteria, cached_status, summary_codes)
                    if sheet is not None
                    else []
                )
                if sheet is not None:
                    used_sheets.add(sheet.title)
                hostname = row["hostname"]
                ip = row["ip_address"]
                os_name = row["target_os"]
                if sheet is not None:
                    _hn, sheet_ip, sheet_os = _host_identity(sheet)
                    ip = ip or sheet_ip
                    os_name = os_name or sheet_os
                    hostname = hostname or _hn
                if not items:
                    continue
                hosts.append(
                    ParsedHost(
                        hostname=hostname,
                        ip_address=ip,
                        target_os=os_name or "Linux",
                        usage=row["usage"],
                        cia_c=row["cia_c"],
                        cia_i=row["cia_i"],
                        cia_a=row["cia_a"],
                        diagnostics=items,
                    )
                )

        for sheet_name in _host_sheet_names(wb):
            if sheet_name in used_sheets:
                continue
            ws = wb[sheet_name]
            items = _parse_host_items(ws, criteria, cached_status, summary_codes)
            if not items:
                continue
            hostname, ip, os_name = _host_identity(ws)
            hosts.append(
                ParsedHost(
                    hostname=hostname or sheet_name,
                    ip_address=ip,
                    target_os=os_name or "Linux",
                    diagnostics=items,
                )
            )

        if not hosts:
            raise ValueError("상세결과보고서에서 호스트 시트를 찾지 못했습니다.")
        return hosts
    finally:
        wb.close()


def _first_item_code(hosts: list[ParsedHost]) -> str:
    for host in hosts:
        for diag in host.diagnostics:
            if diag.code:
                return diag.code.upper()
    return ""


def _os_group_from_hosts(hosts: list[ParsedHost]) -> str:
    code = _first_item_code(hosts)
    if code.startswith("PC-"):
        return "WINDOWS"
    if code.startswith("S-"):
        return "SECURITY"
    if code.startswith("D-"):
        return "DBMS"
    return _export_os_group_key(hosts[0].target_os if hosts else "")


def export_from_detailed_report(path: Path) -> ExportResult:
    logs: list[str] = []
    files: list[str] = []
    if not path.exists() or not path.is_file():
        raise ValueError("상세결과보고서 파일을 찾을 수 없습니다.")

    logs.append(f"[{datetime.now():%H:%M:%S}] 상세결과보고서 읽는 중: {path.name}")
    hosts = parse_detailed_report(path)
    os_type = _os_group_from_hosts(hosts)
    code = _first_item_code(hosts)
    logs.append(
        f"[{datetime.now():%H:%M:%S}] 호스트 {len(hosts)}대, 자산유형 {os_type}, 항목 {code or '-'}…"
    )
    if code and not code.startswith(("U-", "PC-", "S-")):
        logs.append(
            f"[{datetime.now():%H:%M:%S}]   - [안내] 이 자산유형({code})은 UNIX 양식 기준으로 변환되며 "
            f"항목 결과는 N/A로 남을 수 있습니다."
        )

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    logs.append(f"[{datetime.now():%H:%M:%S}] 위험분석 및 평가 보고서 생성 중...")
    out = generate_risk_workbook(hosts, os_type)  # type: ignore[arg-type]
    files.append(out.name)
    logs.append(f"[{datetime.now():%H:%M:%S}]   - 완료: {out.name}")

    logs.append(f"[{datetime.now():%H:%M:%S}] 변환이 완료되었습니다.")
    return ExportResult(files=files, logs=logs)
