from __future__ import annotations

import re
import shutil
import zipfile
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import BarChart, RadarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.graphic import GraphicFrameLocking
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorClientData,
    AnchorMarker,
    SpreadsheetDrawing,
    TwoCellAnchor,
)
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RichTextProperties,
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Chart palette (corporate blue — live cell-linked series)
CHART_FILL = "2E75B6"
CHART_LINE = "1F4E79"
CHART_FILL_ALT = "5B9BD5"
CHART_LINE_ALT = "2F5496"

AREA_VERTEX_LABELS = (
    (18, "① 계정 관리"),
    (19, "② 파일·디렉터리"),
    (20, "③ 서비스 관리"),
    (21, "④ 패치 관리"),
    (22, "⑤ 로그 관리"),
)
AREA_VERTEX_LABELS_DBMS = (
    (18, "① 계정 관리"),
    (19, "② 접근통제"),
    (20, "③ 권한/옵션"),
    (21, "④ 패치/감사"),
    (22, "-"),
)
SECURITY_VERTEX_LABELS = (
    (6, "① 계정 관리"),
    (7, "② 파일·디렉터리"),
    (8, "③ 서비스 관리"),
    (9, "④ 패치 관리"),
    (10, "⑤ 로그 관리"),
)
SECURITY_VERTEX_LABELS_DBMS = (
    (6, "① 계정 관리"),
    (7, "② 접근통제"),
    (8, "③ 권한/옵션"),
    (9, "④ 패치/감사"),
    (10, "-"),
)

from app.config import EXCEL_TEMPLATE_DIR, EXPORT_DIR
from app.models import HostReport
from app.services.stats import classify_host_type

FAIL_FILL = PatternFill(fill_type="solid", fgColor="FFFF0000")
FAIL_FONT = Font(bold=True, color="FFFFFFFF")
PASS_FILL = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
PASS_FONT = Font(bold=True, color="FF000000")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FFFFC000")  # 연주황
WARN_FONT = Font(bold=True, color="FF000000")
NA_FILL = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
NA_FONT = Font(bold=True, color="FF000000")


def _cf_fill(rgb: str) -> PatternFill:
    """Excel CF ignores fills that only set fgColor; set both fg and bg."""
    return PatternFill(patternType="solid", fgColor=rgb, bgColor=rgb)


# CF rules for O-column dropdown (fill+font update together on value change).
CF_FAIL_FILL = _cf_fill("FFFF0000")
CF_PASS_FILL = _cf_fill("FFFFFFFF")
CF_WARN_FILL = _cf_fill("FFFFC000")
CF_NA_FILL = _cf_fill("FFD9D9D9")

_CENTER = Alignment(horizontal="center", vertical="center")
_WARN_LABELS = {"인터뷰", "수동점검", "ERROR", "부분만족"}
RESULT_ENUM_VALUES = ("취약", "N/A", "수동점검", "양호")


@dataclass
class ExportOptions:
    detail: bool = True
    summary: bool = False
    action_plan: bool = False
    unix: bool = True
    dbms: bool = True
    win_server: bool = True
    pc: bool = True


@dataclass
class ExportResult:
    files: list[str]
    logs: list[str]


def _unique_path(directory: Path, base: str, ext: str) -> Path:
    candidate = directory / f"{base}{ext}"
    idx = 1
    while candidate.exists():
        candidate = directory / f"{base}_{idx}{ext}"
        idx += 1
    return candidate


def _status_ko_detail(status: str) -> str:
    """Map scanner status to O-column enum: 취약 / N/A / 수동점검 / 양호."""
    raw = (status or "").strip()
    if raw in RESULT_ENUM_VALUES:
        return raw
    s = raw.lower()
    if s in {"pass", "good", "y"}:
        return "양호"
    if s in {"fail", "vulnerable", "n"}:
        return "취약"
    if s in {"interview", "manual", "error", "인터뷰", "수동점검", "err"}:
        return "수동점검"
    if s in {"n/a", "na", "notapplicable", "not applicable", "not_applicable"}:
        return "N/A"
    return "N/A"


def _status_yn(status: str) -> str:
    s = (status or "").strip().lower()
    if s in {"pass", "good"}:
        return "Y"
    if s in {"fail", "vulnerable"}:
        return "N"
    if s in {"interview", "manual", "error"}:
        return "N/A"
    return "N/A"


def _filter_reports(reports: list[HostReport], options: ExportOptions) -> list[HostReport]:
    selected: list[HostReport] = []
    for report in reports:
        host_type = classify_host_type(report.target_os)
        if host_type == "UNIX/Linux" and options.unix:
            selected.append(report)
        elif host_type == "DBMS" and options.dbms:
            selected.append(report)
        elif host_type == "Windows Server" and options.win_server:
            selected.append(report)
        elif host_type == "개인 PC" and options.pc:
            selected.append(report)
        elif host_type == "기타":
            selected.append(report)
    return selected


def _export_os_group_key(target_os: str) -> str:
    """AIX/Linux/Synology → LINUX, Oracle/DBMS → DBMS, Windows 분리."""
    raw = (target_os or "").strip()
    u = raw.upper()
    if any(m in u for m in ("ORACLE", "DBMS", "MSSQL", "MYSQL", "POSTGRES", "SQL SERVER")):
        return "DBMS"
    if "WINDOW" in u:
        if "SERVER" in u or "서버" in raw:
            return "WINDOWS_SERVER"
        return "WINDOWS"
    unix_markers = (
        "AIX",
        "LINUX",
        "UNIX",
        "SOLARIS",
        "HP-UX",
        "HPUX",
        "CENTOS",
        "RHEL",
        "UBUNTU",
        "SYNOLOGY",
        "DSM",
    )
    if any(m in u for m in unix_markers) or not u:
        return "LINUX"
    return u


def _is_dbms_group(os_type: str) -> bool:
    return (os_type or "").strip().upper() == "DBMS"


def _group_by_os(reports: list[HostReport]) -> dict[str, list[HostReport]]:
    groups: dict[str, list[HostReport]] = {}
    for report in reports:
        key = _export_os_group_key(report.target_os)
        groups.setdefault(key, []).append(report)
    return groups


def _diag_map(report: HostReport) -> dict[str, object]:
    return {d.code.upper(): d for d in report.diagnostics if d.code}


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join("_" if c in r"[]:*?/\\" else c for c in (name or "host"))
    return cleaned[:31] or "host"


def _apply_result_style(cell, result_text: str, *, direct: bool = True) -> None:
    """Style 점검결과 cell.

    direct=False: clear static fill so conditional formatting owns bg+font
    when the user changes the dropdown value.
    """
    cell.alignment = _CENTER
    if not direct:
        cell.fill = PatternFill(fill_type=None)
        cell.font = Font(bold=True, color="FF000000")
        return
    text = (result_text or "").strip()
    if text == "취약" or text == "N":
        cell.fill = FAIL_FILL
        cell.font = FAIL_FONT
    elif text == "양호" or text == "Y":
        cell.fill = PASS_FILL
        cell.font = PASS_FONT
    elif text in _WARN_LABELS:
        cell.fill = WARN_FILL
        cell.font = WARN_FONT
    elif text.upper() == "N/A":
        cell.fill = NA_FILL
        cell.font = NA_FONT


def _clear_chart_cache(chart) -> None:
    """Drop cached series values so Excel refreshes from live cell references."""
    for series in getattr(chart, "series", []) or []:
        val = getattr(series, "val", None)
        if val is not None and getattr(val, "numRef", None) is not None:
            val.numRef.numCache = None
        cat = getattr(series, "cat", None)
        if cat is not None:
            if getattr(cat, "numRef", None) is not None:
                cat.numRef.numCache = None
            if getattr(cat, "strRef", None) is not None:
                cat.strRef.strCache = None


def _pin_chart(
    chart,
    *,
    from_col: int,
    from_row: int,
    to_col: int,
    to_row: int,
) -> None:
    """Bind chart to a fixed TwoCellAnchor (editAs=absolute) that cannot move with cells."""
    anchor = TwoCellAnchor(
        editAs="absolute",
        _from=AnchorMarker(col=from_col, colOff=0, row=from_row, rowOff=0),
        to=AnchorMarker(col=to_col, colOff=0, row=to_row, rowOff=0),
        clientData=AnchorClientData(fLocksWithSheet=True, fPrintsWithSheet=True),
    )
    chart.anchor = anchor
    _clear_chart_cache(chart)


def _install_chart_frame_locks():
    """Patch openpyxl so written chart frames get noMove/noResize locks."""
    original = SpreadsheetDrawing._chart_frame

    def _locked_chart_frame(self, idx):
        frame = original(self, idx)
        frame.nvGraphicFramePr.cNvGraphicFramePr.graphicFrameLocks = GraphicFrameLocking(
            noGrp=True,
            noMove=True,
            noResize=True,
            noChangeAspect=True,
        )
        return frame

    SpreadsheetDrawing._chart_frame = _locked_chart_frame  # type: ignore[method-assign]
    return original


def _restore_chart_frame_locks(original) -> None:
    SpreadsheetDrawing._chart_frame = original  # type: ignore[method-assign]


def _fill_targets_sheet(ws, reports: list[HostReport], sheet_names: list[str]) -> None:
    if ws is None:
        return
    for i, (report, sheet_name) in enumerate(zip(reports, sheet_names)):
        row = 7 + i
        if i > 0:
            for col in range(1, 12):
                src = ws.cell(7, col)
                dst = ws.cell(row, col)
                if src.has_style:
                    dst._style = src._style
        ws.cell(row, 1).value = i + 1
        # B열은 호스트 시트명과 반드시 동일해야 INDIRECT/장비별 공식이 동작한다.
        ws.cell(row, 2).value = sheet_name
        ws.cell(row, 3).value = report.target_os
        ws.cell(row, 4).value = report.ip_address


def _extract_evidence_section(evidence: str, header: str) -> str:
    text = evidence or ""
    if header not in text:
        return ""
    after = text.split(header, 1)[1]
    # Stop at the next [섹션] header if present.
    chunks = after.split("\n[")
    return chunks[0].strip()


def _lookup_guide_comment(
    guidelines: list[dict],
    target_os: str,
    code: str,
    *,
    kind: str,
) -> str:
    """현황 멘트용 가이드. Windows/DBMS 분리, AIX/Synology 등은 Linux 가이드 사용."""
    from app.services.json_parser import guide_os_type

    guide_os = guide_os_type(target_os, code)
    for g in guidelines or []:
        if (
            str(g.get("os_type", "")).lower() == guide_os.lower()
            and str(g.get("code", "")).lower() == (code or "").lower()
        ):
            return str(g.get(kind) or "").strip()
    return ""


_CRITERIA_CACHE: dict[str, dict[str, str]] | None = None


def _load_criteria_map() -> dict[str, dict[str, str]]:
    """템플릿 기준 DB → 코드별 양호/취약 기준 문구 (UNIX + DBMS)."""
    global _CRITERIA_CACHE
    if _CRITERIA_CACHE is not None:
        return _CRITERIA_CACHE
    mapping: dict[str, dict[str, str]] = {}
    for name in (
        "UNIX_서버_취약점진단_상세결과보고서.xlsx",
        "DBMS_서버_취약점진단_상세결과보고서.xlsx",
    ):
        path = EXCEL_TEMPLATE_DIR / name
        if not path.exists():
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if "기준 DB" in wb.sheetnames:
                for row in wb["기준 DB"].iter_rows(min_row=2, max_col=4, values_only=True):
                    code = str(row[0] or "").strip().upper()
                    if not code:
                        continue
                    mapping[code] = {
                        "title": str(row[1] or "").strip(),
                        "pass": str(row[2] or "").strip(),
                        "fail": str(row[3] or "").strip(),
                    }
            wb.close()
        except Exception:
            continue
    _CRITERIA_CACHE = mapping
    return mapping


def _extract_criteria_body(block: str) -> str:
    text = (block or "").strip()
    if not text:
        return ""
    if "■ 기준" in text:
        text = text.split("■ 기준", 1)[1]
        if ":" in text:
            text = text.split(":", 1)[1]
    if "■ 현황" in text:
        text = text.split("■ 현황", 1)[0]
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s or s.startswith("※"):
            continue
        lines.append(s)
    return " ".join(lines).strip()


def _criteria_to_status_comment(block: str, *, kind: str) -> str:
    """기준 문구를 ■ 현황에 붙일 형식적 문장으로 변환."""
    ending = "양호함" if kind == "pass" else "취약함"
    body = _extract_criteria_body(block)
    if not body:
        return ""
    body = body.replace("존재하지 경우", "존재하는 경우")
    body = re.sub(r'\s*경우\s*["“”]?양호["“”]?\s*$', "", body)
    body = re.sub(r'\s*경우\s*["“”]?취약["“”]?\s*$', "", body)
    body = body.strip().rstrip(",").strip()
    if not body:
        return ""
    replacements = (
        ("하지 않는", "하지 않아"),
        ("되지 않은", "되지 않아"),
        ("되지 않는", "되지 않아"),
        ("존재하지 않는", "존재하지 않아"),
        ("존재하지 않은", "존재하지 않아"),
        ("없는", "없어"),
        ("있는", "있어"),
        ("하는", "하여"),
        ("되는", "되어"),
        ("된", "되어"),
        ("한", "하여"),
        ("인", "이므로"),
    )
    converted = False
    for old, new in replacements:
        if body.endswith(old):
            body = body[: -len(old)] + new
            converted = True
            break
    if body.endswith(ending):
        return body
    if converted or body.endswith(("하여", "되어", "않아", "있어", "없어", "이므로")):
        return f"{body} {ending}"
    return f"{body}이므로 {ending}"


def _to_ham_status_comment(text: str, *, ending: str) -> str:
    """Normalize formal comments to …양호함 / …취약함 endings."""
    t = (text or "").strip().rstrip(".")
    replacements = (
        ("양호합니다", "양호함"),
        ("취약합니다", "취약함"),
        ("안전합니다", "양호함"),
        ("양호하다", "양호함"),
        ("취약하다", "취약함"),
    )
    for old, new in replacements:
        if t.endswith(old):
            t = t[: -len(old)] + new
            break
    if ending == "양호함" and not t.endswith("양호함"):
        if not t:
            return ""
        if t.endswith("지 않습니다") or t.endswith("지 않았습니다"):
            t = t.rsplit("지 않", 1)[0] + "되어 양호함"
        elif t.endswith("있습니다"):
            t = t[: -len("있습니다")] + "있어 양호함"
        elif t.endswith("습니다"):
            t = t[: -len("습니다")] + " 양호함"
        elif t.endswith("다"):
            t = t[:-1] + " 양호함"
        else:
            t = t.rstrip() + " 양호함"
    if ending == "취약함" and not t.endswith("취약함"):
        if not t:
            return ""
        if t.endswith("지 않습니다") or t.endswith("지 않았습니다"):
            t = t.rsplit("지 않", 1)[0] + "지 않아 취약함"
        elif t.endswith("있습니다"):
            t = t[: -len("있습니다")] + "있어 취약함"
        elif t.endswith("습니다"):
            t = t[: -len("습니다")] + " 취약함"
        elif t.endswith("다"):
            t = t[:-1] + " 취약함"
        else:
            t = t.rstrip() + " 취약함"
    return t


def _is_korean_comment(text: str) -> bool:
    t = (text or "").strip()
    if not t:
        return False
    return sum(1 for ch in t if "가" <= ch <= "힣") >= 4


def _item_title_for_status(diag, guidelines: list[dict], target_os: str) -> str:
    code = (getattr(diag, "code", "") or "").strip().upper()
    criteria = _load_criteria_map().get(code) or {}
    if criteria.get("title"):
        return criteria["title"]
    from app.services.json_parser import guide_os_type

    guide_os = guide_os_type(target_os, code)
    for g in guidelines or []:
        if (
            str(g.get("os_type", "")).lower() == guide_os.lower()
            and str(g.get("code", "")).lower() == code.lower()
        ):
            t = str(g.get("title") or "").strip()
            if t:
                return t
    return code or "해당 항목"


def _build_pass_status_comment(
    diag,
    *,
    guidelines: list[dict],
    target_os: str,
) -> str:
    """U열: ■ 현황 양호 멘트 — 기준과 같은 일률 문구."""
    code = (getattr(diag, "code", "") or "").strip().upper()
    # 1) 기준 DB 양호 기준을 현황처럼 되풀이 (일률)
    criteria = _load_criteria_map().get(code) or {}
    text = _criteria_to_status_comment(criteria.get("pass", ""), kind="pass")
    if _is_korean_comment(text):
        return text
    # 2) 가이드 한국어 (기준 DB 없을 때만)
    raw = _lookup_guide_comment(guidelines, target_os, code, kind="pass_comment")
    text = _to_ham_status_comment(raw, ending="양호함") if raw else ""
    if _is_korean_comment(text):
        return text
    title = _item_title_for_status(diag, guidelines, target_os)
    return f"{title} 설정이 기준에 맞게 적용되어 양호함"


def _build_fail_status_comment(
    diag,
    *,
    guidelines: list[dict],
    target_os: str,
) -> str:
    """V열: ■ 현황 취약 멘트 — 기준과 같은 일률 문구."""
    code = (getattr(diag, "code", "") or "").strip().upper()
    # 1) 기준 DB 취약 기준을 현황처럼 되풀이 (일률)
    criteria = _load_criteria_map().get(code) or {}
    text = _criteria_to_status_comment(criteria.get("fail", ""), kind="fail")
    if _is_korean_comment(text):
        return text
    # 2) 가이드 한국어 (기준 DB 없을 때만)
    raw = _lookup_guide_comment(guidelines, target_os, code, kind="fail_comment")
    text = _to_ham_status_comment(raw, ending="취약함") if raw else ""
    if _is_korean_comment(text):
        return text
    title = _item_title_for_status(diag, guidelines, target_os)
    return f"{title} 기준 미달 설정이 적용되어 취약함"


def _inspection_status_comment(
    diag,
    *,
    guidelines: list[dict],
    target_os: str,
) -> str:
    """단일 현황 멘트 (현재 결과에 맞는 형식 문장)."""
    result_ko = _status_ko_detail(getattr(diag, "status", "") or "")
    if result_ko == "양호":
        return _build_pass_status_comment(diag, guidelines=guidelines, target_os=target_os)
    if result_ko == "취약":
        return _build_fail_status_comment(diag, guidelines=guidelines, target_os=target_os)
    return ""


_OP_FONT_RED = InlineFont(b=False, sz=10, color="FFFF0000")
_OP_FONT_RED_BODY = InlineFont(b=False, sz=10, color="FFFF0000")
_OP_FONT_NORMAL = InlineFont(b=False, sz=10, color="FF000000")


def _evidence_preamble(evidence: str) -> str:
    """Text before the first [섹션] header — often the fail/pass banner."""
    text = (evidence or "").strip()
    if not text:
        return ""
    if "\n[" in text:
        return text.split("\n[", 1)[0].strip()
    if text.startswith("["):
        return ""
    return text


def _operation_status_text(diag) -> str | CellRichText:
    """운영현황(Q열): [취약점 현황] → [설정값 현황] → [설정값 상세] (양호 포함 모든 결과).

    [취약점 현황] 헤더+본문은 전부 빨간색.
    """
    evidence = getattr(diag, "evidence", None) or ""
    processed = (getattr(diag, "processed_config", None) or "").strip()
    vulnerable = (getattr(diag, "vulnerable_config", None) or "").strip()
    raw = (getattr(diag, "raw_config", None) or "").strip()

    # Legacy rows (uploaded before raw/vulnerable columns): recover from evidence.
    if not processed:
        processed = _extract_evidence_section(evidence, "[검출된 설정값 (ProcessedConfig)]")
    if not vulnerable:
        vulnerable = _extract_evidence_section(evidence, "[취약 근거]")
    if not vulnerable:
        vulnerable = _extract_evidence_section(evidence, "[취약한 설정 분석 내용]")
    if not raw:
        raw = _extract_evidence_section(evidence, "[진단 로그 / 설정 원본 (RawConfig)]")

    status = (getattr(diag, "status", "") or "").strip().lower()

    # Fail with empty VulnerableConfig: use leading evidence sentence as 취약점 현황.
    if not vulnerable and status in {"fail", "vulnerable"}:
        preamble = _evidence_preamble(evidence)
        if preamble and preamble.upper() != "N/A" and "양호" not in preamble:
            vulnerable = preamble

    # Keep Excel cells usable — Synology RawConfig can be large.
    if len(raw) > 8000:
        raw = raw[:8000] + f"\n... (truncated, total {len(raw)} chars)"

    blocks: list[TextBlock] = []

    def _append_section(
        *,
        header_font: InlineFont,
        body_font: InlineFont,
        header: str,
        body: str,
    ) -> None:
        if not body:
            return
        prefix = "\n\n" if blocks else ""
        blocks.append(TextBlock(header_font, f"{prefix}[{header}]"))
        blocks.append(TextBlock(body_font, f"\n{body}"))

    # 취약점 현황: header + body all red
    _append_section(
        header_font=_OP_FONT_RED,
        body_font=_OP_FONT_RED_BODY,
        header="취약점 현황",
        body=vulnerable,
    )
    _append_section(
        header_font=_OP_FONT_NORMAL,
        body_font=_OP_FONT_NORMAL,
        header="설정값 현황",
        body=processed,
    )
    _append_section(
        header_font=_OP_FONT_NORMAL,
        body_font=_OP_FONT_NORMAL,
        header="설정값 상세",
        body=raw,
    )

    if not blocks:
        return ""
    return CellRichText(*blocks)


def _scale_item_row_heights(ws, start_row: int, end_row: int, code_col: int, factor: float = 1.5) -> None:
    default_h = ws.sheet_format.defaultRowHeight or 15.0
    for row in range(start_row, end_row + 1):
        if not str(ws.cell(row, code_col).value or "").strip():
            continue
        current = ws.row_dimensions[row].height
        base = float(current) if current else float(default_h)
        ws.row_dimensions[row].height = base * factor


def _fill_host_detail_sheet(
    ws,
    report: HostReport,
    *,
    code_col: int,
    result_col: int,
    result_mapper,
    evidence_col: int | None,
    op_col: int | None,
    start_row: int,
    end_row: int,
    guidelines: list[dict] | None = None,
    short_inspection_status: bool = False,
    pass_comment_col: int | None = None,
    fail_comment_col: int | None = None,
) -> None:
    mapping = _diag_map(report)
    guides = guidelines or []
    u_col = pass_comment_col or (evidence_col if short_inspection_status else None)
    v_col = fail_comment_col
    for row in range(start_row, end_row + 1):
        code = str(ws.cell(row, code_col).value or "").strip()
        if not code:
            continue
        result_cell = ws.cell(row, result_col)
        diag = mapping.get(code.upper())
        if diag is None:
            result_cell.value = "N/A"
            _apply_result_style(result_cell, "N/A", direct=not short_inspection_status)
            if short_inspection_status:
                if u_col:
                    ws.cell(row, u_col).value = ""
                if v_col:
                    ws.cell(row, v_col).value = ""
            elif evidence_col:
                ws.cell(row, evidence_col).value = ""
            if op_col:
                ws.cell(row, op_col).value = ""
            continue
        result_text = result_mapper(diag.status)
        result_cell.value = result_text
        # Detail report: CF owns fill+font so dropdown changes update background too.
        _apply_result_style(result_cell, result_text, direct=not short_inspection_status)
        if short_inspection_status and u_col:
            # U=양호 형식 멘트, V=취약 형식 멘트 (P수식이 O에 따라 선택)
            ws.cell(row, u_col).value = _build_pass_status_comment(
                diag, guidelines=guides, target_os=report.target_os
            )
            ws.cell(row, u_col).alignment = Alignment(wrap_text=True, vertical="top")
            if v_col:
                ws.cell(row, v_col).value = _build_fail_status_comment(
                    diag, guidelines=guides, target_os=report.target_os
                )
                ws.cell(row, v_col).alignment = Alignment(wrap_text=True, vertical="top")
        elif evidence_col:
            ws.cell(row, evidence_col).value = diag.evidence or ""
            ws.cell(row, evidence_col).alignment = Alignment(wrap_text=True, vertical="top")
        if op_col:
            if short_inspection_status:
                ws.cell(row, op_col).value = _operation_status_text(diag)
            else:
                ws.cell(row, op_col).value = diag.processed_config or diag.err_msg or ""
            ws.cell(row, op_col).alignment = Alignment(wrap_text=True, vertical="top")

    _scale_item_row_heights(ws, start_row, end_row, code_col, factor=1.5)

def _reapply_detail_conditional_formatting(ws) -> None:
    """O열 값 변경 시 글씨·배경색이 따라가도록 CF 적용 (드롭다운 연동)."""
    try:
        ws.conditional_formatting._cf_rules.clear()
    except Exception:
        pass

    # Priority: first match wins (stopIfTrue).
    # Use CF_* fills (fg+bg) — Excel ignores CF fills that only set fgColor.
    ws.conditional_formatting.add(
        "O28:O94",
        CellIsRule(
            operator="equal",
            formula=['"취약"'],
            fill=CF_FAIL_FILL,
            font=FAIL_FONT,
            stopIfTrue=True,
        ),
    )
    ws.conditional_formatting.add(
        "O28:O94",
        CellIsRule(
            operator="equal",
            formula=['"수동점검"'],
            fill=CF_WARN_FILL,
            font=WARN_FONT,
            stopIfTrue=True,
        ),
    )
    ws.conditional_formatting.add(
        "O28:O94",
        CellIsRule(
            operator="equal",
            formula=['"양호"'],
            fill=CF_PASS_FILL,
            font=PASS_FONT,
            stopIfTrue=True,
        ),
    )
    ws.conditional_formatting.add(
        "O28:O94",
        CellIsRule(
            operator="equal",
            formula=['"N/A"'],
            fill=CF_NA_FILL,
            font=NA_FONT,
            stopIfTrue=True,
        ),
    )


def _add_result_enum_dropdown(ws, start_row: int = 28, end_row: int = 94) -> None:
    """O열 점검결과를 취약/N/A/수동점검/양호 중 선택하도록 제한."""
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(RESULT_ENUM_VALUES) + '"',
        allow_blank=True,
        showDropDown=False,  # False = show dropdown arrow in Excel
        showErrorMessage=True,
        errorTitle="점검결과",
        error="취약, N/A, 수동점검, 양호 중에서 선택하세요.",
        promptTitle="점검결과",
        prompt="취약 / N/A / 수동점검 / 양호",
    )
    dv.add(f"O{start_row}:O{end_row}")
    ws.add_data_validation(dv)


def _is_dbms_detail_sheet(ws, code_col: int = 3) -> bool:
    code = str(ws.cell(28, code_col).value or "").strip().upper()
    return code.startswith("D-")


def _fix_host_stat_formulas(ws) -> None:
    """Count 취약 only via COUNTIF; interview/manual/error go to 해당없음 bucket."""
    if _is_dbms_detail_sheet(ws):
        specs = [
            (18, "O28:O36"),
            (19, "O37:O45"),
            (20, "O46:O51"),
            (21, "O52:O53"),
        ]
        # Row 22 unused for DBMS — keep zeros from template.
        for row, rng in specs:
            ws.cell(row, 6).value = f'=COUNTIF({rng},"양호")'
            ws.cell(row, 7).value = f'=COUNTIF({rng},"취약")'
            ws.cell(row, 8).value = f'=COUNTIF({rng},"부분만족")'
            ws.cell(row, 9).value = (
                f'=COUNTIF({rng},"N/A")+COUNTIF({rng},"인터뷰")'
                f'+COUNTIF({rng},"수동점검")+COUNTIF({rng},"ERROR")'
            )
        return
    specs = [
        (18, "O28:O40"),
        (19, "O41:O60"),
        (20, "O61:O90"),
        (21, "O91"),
        (22, "O92:O94"),
    ]
    for row, rng in specs:
        ws.cell(row, 6).value = f'=COUNTIF({rng},"양호")'
        ws.cell(row, 7).value = f'=COUNTIF({rng},"취약")'
        ws.cell(row, 8).value = f'=COUNTIF({rng},"부분만족")'
        ws.cell(row, 9).value = (
            f'=COUNTIF({rng},"N/A")+COUNTIF({rng},"인터뷰")'
            f'+COUNTIF({rng},"수동점검")+COUNTIF({rng},"ERROR")'
        )


def _fix_detail_status_formulas(ws, start_row: int = 28, end_row: int = 94) -> None:
    """P열: O값에 따라 양호→기준C&U, 취약→기준D&V. 수동점검/N/A는 현황 빈칸."""
    for row in range(start_row, end_row + 1):
        # 기준 DB rows: U-01 at row 2 → item row 28 maps to db row (row-26)
        db_row = row - 26
        ws.cell(row, 16).value = (
            f'=IF(O{row}="양호",\'기준 DB\'!C{db_row}&U{row},'
            f'IF(O{row}="취약",\'기준 DB\'!D{db_row}&V{row},'
            f'IF(OR(O{row}="N/A",O{row}="수동점검"),\'기준 DB\'!E{db_row},"-")))'
        )


def _percent_data_labels(*, position: str | None = "outEnd") -> DataLabelList:
    labels = DataLabelList()
    labels.showVal = True
    labels.showCatName = False
    labels.showSerName = False
    labels.showPercent = False
    labels.showLegendKey = False
    labels.numFmt = "0%"
    if position:
        labels.dLblPos = position
    return labels


def _style_series(series, *, fill: str = CHART_FILL, line: str = CHART_LINE) -> None:
    series.graphicalProperties = GraphicalProperties(
        solidFill=fill,
        ln=LineProperties(solidFill=line, w=25000),
    )


def _write_vertex_labels(ws, labels: tuple[tuple[int, str], ...], col: int) -> None:
    """Helper column used as radar/bar categories (numbered vertex names)."""
    for row, text in labels:
        cell = ws.cell(row, col)
        cell.value = text
        cell.font = Font(name="맑은 고딕", size=9, bold=True, color="FF1F4E79")


def _category_formula(sheet_title: str, col_letter: str, start_row: int, end_row: int) -> str:
    safe = sheet_title.replace("'", "''")
    return f"'{safe}'!${col_letter}${start_row}:${col_letter}${end_row}"


def _bind_str_categories(chart, formula: str) -> None:
    """Force text category axis (openpyxl defaults to numRef, which hides X labels)."""
    for series in chart.series:
        series.cat = AxDataSource(strRef=StrRef(f=formula))


def _rotate_axis_labels(axis, degrees: float = -45.0) -> None:
    # OOXML rotation unit: 1/60000 of a degree.
    rot = int(degrees * 60000)
    axis.txPr = RichText(
        bodyPr=RichTextProperties(rot=rot, anchor="ctr"),
        p=[
            Paragraph(
                pPr=ParagraphProperties(defRPr=CharacterProperties(sz=900)),
                endParaRPr=CharacterProperties(sz=900),
            )
        ],
    )


def _build_radar_chart(
    *,
    title: str | None,
    cats: Reference,
    data: Reference,
    cat_formula: str,
    from_col: int,
    from_row: int,
    to_col: int,
    to_row: int,
) -> RadarChart:
    chart = RadarChart()
    chart.type = "filled"
    chart.style = 10
    chart.title = title
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _bind_str_categories(chart, cat_formula)
    # Hide series legend — category names already sit on each vertex axis.
    chart.legend = None
    chart.dataLabels = _percent_data_labels(position=None)
    for series in chart.series:
        _style_series(series, fill=CHART_FILL_ALT, line=CHART_LINE)
        series.dLbls = _percent_data_labels(position=None)
    try:
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
        # Keep concentric grid; tick text is stripped in _hide_radar_value_axis_labels().
        chart.y_axis.numFmt = "0%"
        chart.y_axis.majorTickMark = "none"
        chart.y_axis.minorTickMark = "none"
        chart.y_axis.delete = False
        chart.x_axis.delete = False
    except Exception:
        pass
    chart.width = 12
    chart.height = 10
    _pin_chart(chart, from_col=from_col, from_row=from_row, to_col=to_col, to_row=to_row)
    return chart


def _hide_radar_value_axis_labels(xlsx_path: Path) -> None:
    """openpyxl cannot emit tickLblPos=none; patch radar valAx after save."""

    def _patch_val_ax(match: re.Match[str]) -> str:
        block = match.group(0)
        # Preserve whatever prefix the chart XML uses (c: or none).
        prefix = "c:" if "<c:axPos" in block or block.startswith("<c:") else ""
        tick_tag = f"<{prefix}tickLblPos val=\"none\"/>"
        if re.search(rf"<{prefix}tickLblPos\s+val=\"none\"", block):
            return block
        if f"<{prefix}tickLblPos" in block or "<tickLblPos" in block:
            return re.sub(r"<[^>]*tickLblPos[^/]*/>", tick_tag, block)
        return re.sub(
            rf"(<{prefix}axPos[^/]*/>)",
            rf"\1{tick_tag}",
            block,
            count=1,
        )

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        files = {info.filename: (info, zin.read(info.filename)) for info in zin.infolist()}

    changed = False
    for name, (info, data) in list(files.items()):
        if not (name.startswith("xl/charts/") and name.endswith(".xml")):
            continue
        text = data.decode("utf-8")
        if "<radarChart>" not in text and "<c:radarChart>" not in text:
            continue
        patched = re.sub(
            r"<(?:c:)?valAx>.*?</(?:c:)?valAx>",
            _patch_val_ax,
            text,
            flags=re.S,
        )
        if patched != text:
            files[name] = (info, patched.encode("utf-8"))
            changed = True

    if not changed:
        return

    tmp = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")
    with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name, (info, data) in files.items():
            zout.writestr(info, data)
    tmp.replace(xlsx_path)


def _build_bar_chart(
    *,
    title: str | None,
    cats: Reference,
    data: Reference,
    cat_formula: str,
    from_col: int,
    from_row: int,
    to_col: int,
    to_row: int,
    rotate_labels: bool = True,
) -> BarChart:
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _bind_str_categories(chart, cat_formula)
    # Single series — legend only adds clutter under the x-axis labels.
    chart.legend = None
    chart.dataLabels = _percent_data_labels(position="outEnd")
    for series in chart.series:
        _style_series(series, fill=CHART_FILL, line=CHART_LINE)
        series.dLbls = _percent_data_labels(position="outEnd")
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.numFmt = "0%"
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    if rotate_labels:
        _rotate_axis_labels(chart.x_axis, -45.0)
    chart.shape = 4
    chart.width = 12
    chart.height = 10
    _pin_chart(chart, from_col=from_col, from_row=from_row, to_col=to_col, to_row=to_row)
    return chart


def _add_host_area_charts(ws) -> None:
    """Recreate live, pinned area charts — radar at column P, bar at column Q."""
    ws._charts = []
    # AA열: 꼭짓점/축 라벨 / L열: 보안수준(비율)
    labels = AREA_VERTEX_LABELS_DBMS if _is_dbms_detail_sheet(ws) else AREA_VERTEX_LABELS
    _write_vertex_labels(ws, labels, col=27)
    if not ws.cell(17, 12).value:
        ws.cell(17, 12).value = "보안수준"
    ws.cell(17, 27).value = "점검 영역"

    # Give P/Q enough width so each chart has its own column lane.
    ws.column_dimensions["P"].width = 42
    ws.column_dimensions["Q"].width = 42

    cats = Reference(ws, min_col=27, min_row=18, max_row=22)
    data = Reference(ws, min_col=12, min_row=17, max_row=22)
    cat_formula = _category_formula(ws.title, "AA", 18, 22)

    # P=15, Q=16 (0-based anchor indices); from_row=6 → Excel 7행
    radar = _build_radar_chart(
        title=None,
        cats=cats,
        data=data,
        cat_formula=cat_formula,
        from_col=15,
        from_row=6,
        to_col=16,
        to_row=22,
    )
    ws.add_chart(radar)

    bar = _build_bar_chart(
        title=None,
        cats=cats,
        data=data,
        cat_formula=cat_formula,
        from_col=16,
        from_row=6,
        to_col=17,
        to_row=22,
    )
    ws.add_chart(bar)


def _fill_summary_sheet(ws, host_count: int) -> None:
    if ws is None or host_count < 1:
        return
    # Template has host results starting at column L (12).
    for i in range(host_count):
        col = 12 + i
        letter = get_column_letter(col)
        if i > 0:
            ws.column_dimensions[letter].width = ws.column_dimensions["L"].width or 12
            for row in range(1, 92):
                src = ws.cell(row, 12)
                dst = ws.cell(row, col)
                if src.has_style:
                    dst._style = copy(src._style)

        ws.cell(4, col).value = (
            '=HYPERLINK("[" & MID(CELL("filename"),SEARCH("[",CELL("filename"))+1, '
            'SEARCH("]",CELL("filename"))-SEARCH("[",CELL("filename"))-1) & "]\'" & '
            'INDIRECT(ADDRESS(COLUMN()-5,2,1,TRUE,"점검대상")) & "\'!A1",'
            'INDIRECT(ADDRESS(COLUMN()-5,2,1,TRUE,"점검대상")))'
        )
        for row in range(6, 73):
            ws.cell(row, col).value = (
                f'=IFERROR(INDIRECT(ADDRESS(ROW()+22,15,1,TRUE,{letter}$4)),"")'
            )
        ws.cell(73, col).value = (
            f'=COUNTIF({letter}6:{letter}72,"양호")+COUNTIF({letter}6:{letter}72,"취약")'
            f'+COUNTIF({letter}6:{letter}72,"부분만족")+COUNTIF({letter}6:{letter}72,"N/A")'
        )
        for row in range(75, 79):
            ws.cell(row, col).value = f"=COUNTIF({letter}$6:{letter}$72,$J{row})"
        for row in range(80, 92):
            # Severity blocks: 상@80, 중@84, 하@88 (template I-column anchors).
            if row <= 83:
                anchor_row = 80
            elif row <= 87:
                anchor_row = 84
            else:
                anchor_row = 88
            ws.cell(row, col).value = (
                f"=COUNTIFS($D$6:$D$72,$I${anchor_row},{letter}$6:{letter}$72,$J{row})"
            )

    last_letter = get_column_letter(11 + host_count)
    for row in range(75, 79):
        ws.cell(row, 11).value = f"=SUM(L{row}:{last_letter}{row})"
    for row in range(80, 92):
        ws.cell(row, 11).value = f"=SUM(L{row}:{last_letter}{row})"


def _fill_security_level_sheet(ws, reports: list[HostReport], os_type: str) -> None:
    if ws is None or not reports:
        return

    host_count = len(reports)
    ws["A32"] = "▣ 장비별 보안수준 그래프"

    if host_count > 1:
        # Push "계" row down; openpyxl keeps old merge on A37 so fix merges after.
        try:
            ws.unmerge_cells("A37:B37")
        except Exception:
            pass
        ws.insert_rows(37, host_count - 1)
        for r in range(37, 36 + host_count):
            for c in range(1, 12):
                src = ws.cell(36, c)
                dst = ws.cell(r, c)
                if src.has_style:
                    dst._style = copy(src._style)

    for i in range(host_count):
        r = 36 + i
        ws.cell(r, 1).value = (os_type or "UNIX").upper()
        ws.cell(r, 2).value = '=INDIRECT(ADDRESS(ROW()-29,2,1,TRUE,"점검대상"))'
        for c in range(3, 11):
            ws.cell(r, c).value = f'=IFERROR(INDIRECT(ADDRESS(23,COLUMN()+1,1,TRUE,$B{r})),0)'
        ws.cell(r, 11).value = f'=IF(I{r}=0,"N/A",((I{r}-J{r})/I{r})*1)'

    total_row = 36 + host_count
    # Ensure 계 merge sits on the total row.
    for merged in list(ws.merged_cells.ranges):
        if str(merged).startswith("A37:"):
            try:
                ws.unmerge_cells(str(merged))
            except Exception:
                pass
    try:
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    except Exception:
        pass

    ws.cell(total_row, 1).value = "계"
    for c in range(3, 11):
        letter = get_column_letter(c)
        ws.cell(total_row, c).value = f"=SUM({letter}36:{letter}{total_row - 1})"
    ws.cell(total_row, 11).value = (
        f'=SUM(K36:K{total_row - 1})/'
        f'(COUNTA(K36:K{total_row - 1})-COUNTIF(K36:K{total_row - 1},"N/A"))'
    )
    ws["K11"] = f"=K{total_row}"

    # Rebuild all charts: live cell refs, percent labels, pinned anchors.
    ws._charts = []
    sec_labels = SECURITY_VERTEX_LABELS_DBMS if _is_dbms_group(os_type) else SECURITY_VERTEX_LABELS
    _write_vertex_labels(ws, sec_labels, col=13)  # M열
    if not ws.cell(5, 11).value:
        ws.cell(5, 11).value = "보안수준"
    ws.cell(5, 13).value = "점검 영역"

    area_cats = Reference(ws, min_col=13, min_row=6, max_row=10)
    area_data = Reference(ws, min_col=11, min_row=5, max_row=10)
    area_cat_formula = _category_formula(ws.title, "M", 6, 10)

    # Left radar / right bar — matches 영역별 보안수준 그래프 section layout.
    radar = _build_radar_chart(
        title=None,
        cats=area_cats,
        data=area_data,
        cat_formula=area_cat_formula,
        from_col=0,
        from_row=12,
        to_col=5,
        to_row=30,
    )
    ws.add_chart(radar)

    area_bar = _build_bar_chart(
        title=None,
        cats=area_cats,
        data=area_data,
        cat_formula=area_cat_formula,
        from_col=5,
        from_row=12,
        to_col=11,
        to_row=30,
    )
    ws.add_chart(area_bar)

    last_host_row = total_row - 1
    # Ensure series header for 장비별 chart.
    if not ws.cell(35, 11).value:
        ws.cell(35, 11).value = "보안수준"
    host_cats = Reference(ws, min_col=2, min_row=36, max_row=last_host_row)
    host_data = Reference(ws, min_col=11, min_row=35, max_row=last_host_row)
    host_cat_formula = _category_formula(ws.title, "B", 36, last_host_row)
    host_bar = _build_bar_chart(
        title="장비별 보안수준 (%)",
        cats=host_cats,
        data=host_data,
        cat_formula=host_cat_formula,
        from_col=0,
        from_row=total_row + 1,
        to_col=11,
        to_row=total_row + 16,
        rotate_labels=True,
    )
    host_bar.width = 18
    host_bar.height = 10
    ws.add_chart(host_bar)


def _retarget_security_3d_refs(ws, first_sheet: str) -> None:
    """Point area rollups at host sheets sandwiched before 기준 DB."""
    old = "'sample:기준 DB'"
    new = f"'{first_sheet}:기준 DB'"
    for row in ws.iter_rows(min_row=6, max_row=10, min_col=3, max_col=11):
        for cell in row:
            if isinstance(cell.value, str) and old in cell.value:
                cell.value = cell.value.replace(old, new)


def _move_sheets_before(wb, sheet_names: list[str], before_title: str) -> None:
    if before_title not in wb.sheetnames:
        return
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        # Move sheet to sit immediately before 기준 DB (or current before_title).
        before_idx = wb.sheetnames.index(before_title)
        cur_idx = wb.sheetnames.index(name)
        offset = before_idx - cur_idx
        if offset != 0:
            wb.move_sheet(name, offset=offset)


def generate_detailed_report(
    reports: list[HostReport],
    os_type: str,
    guidelines: list[dict] | None = None,
) -> Path:
    if _is_dbms_group(os_type):
        template = EXCEL_TEMPLATE_DIR / "DBMS_서버_취약점진단_상세결과보고서.xlsx"
    else:
        template = EXCEL_TEMPLATE_DIR / "UNIX_서버_취약점진단_상세결과보고서.xlsx"
    if not template.exists():
        raise FileNotFoundError(f"상세결과 템플릿 없음: {template.name}")

    clean_os = os_type.replace(" ", "_")
    out = _unique_path(
        EXPORT_DIR,
        f"{clean_os}_서버_취약점진단_상세결과보고서_{datetime.now():%Y%m%d}",
        ".xlsx",
    )
    shutil.copy2(template, out)
    wb = load_workbook(out)

    if "표지" in wb.sheetnames:
        wb["표지"]["H19"] = datetime.now().strftime("%Y.%m.")

    # Unique sheet names (Excel 31-char limit), stable for 점검대상/INDIRECT.
    sheet_names: list[str] = []
    used: set[str] = set()
    for report in reports:
        base = _safe_sheet_name(report.hostname)
        name = base
        n = 2
        while name.lower() in used or name in wb.sheetnames and name != "sample":
            suffix = f"_{n}"
            name = _safe_sheet_name(base[: 31 - len(suffix)] + suffix)
            n += 1
        used.add(name.lower())
        sheet_names.append(name)

    if "점검대상" in wb.sheetnames:
        _fill_targets_sheet(wb["점검대상"], reports, sheet_names)

    if "요약 통계" in wb.sheetnames:
        _fill_summary_sheet(wb["요약 통계"], len(reports))

    sample_name = "sample" if "sample" in wb.sheetnames else None
    created: list[str] = []
    detail_end = 53 if _is_dbms_group(os_type) else 94
    if sample_name:
        for report, sheet_name in zip(reports, sheet_names):
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.copy_worksheet(wb[sample_name])
            ws.title = sheet_name
            created.append(sheet_name)
            for coord, value in (
                ("C6", report.hostname),
                ("C7", report.ip_address),
                ("D4", report.hostname),
                ("D5", report.ip_address),
            ):
                try:
                    ws[coord] = value
                except Exception:
                    pass
            _fill_host_detail_sheet(
                ws,
                report,
                code_col=3,
                result_col=15,
                result_mapper=_status_ko_detail,
                evidence_col=21,
                op_col=17,
                start_row=28,
                end_row=detail_end,
                guidelines=guidelines,
                short_inspection_status=True,
                pass_comment_col=21,  # U: 양호 형식 멘트
                fail_comment_col=22,  # V: 취약 형식 멘트
            )
            _fix_host_stat_formulas(ws)
            _fix_detail_status_formulas(ws, start_row=28, end_row=detail_end)
            _reapply_detail_conditional_formatting(ws)
            _add_result_enum_dropdown(ws, start_row=28, end_row=detail_end)
            _add_host_area_charts(ws)

        try:
            del wb[sample_name]
        except Exception:
            pass

    if created:
        _move_sheets_before(wb, created, "기준 DB")
        if "보안수준 통계" in wb.sheetnames:
            _retarget_security_3d_refs(wb["보안수준 통계"], created[0])
            _fill_security_level_sheet(wb["보안수준 통계"], reports, os_type)

    original_frame = _install_chart_frame_locks()
    try:
        wb.save(out)
    finally:
        _restore_chart_frame_locks(original_frame)
        wb.close()
    _hide_radar_value_axis_labels(out)
    return out


def generate_risk_workbook(reports: list[HostReport], os_type: str, template_name: str, out_prefix: str) -> Path:
    template = EXCEL_TEMPLATE_DIR / template_name
    if not template.exists():
        raise FileNotFoundError(f"템플릿 없음: {template_name}")

    clean_os = os_type.replace(" ", "_")
    out = _unique_path(EXPORT_DIR, f"{clean_os}_{out_prefix}_{datetime.now():%Y%m%d}", ".xlsm")
    shutil.copy2(template, out)
    wb = load_workbook(out, keep_vba=True)

    if "표지" in wb.sheetnames:
        try:
            wb["표지"]["I21"] = datetime.now().strftime("%Y.%m.")
        except Exception:
            pass

    sheet_names: list[str] = []
    used: set[str] = set()
    for report in reports:
        base = _safe_sheet_name(report.hostname)
        name = base
        n = 2
        while name.lower() in used or name in wb.sheetnames and name != "sample":
            suffix = f"_{n}"
            name = _safe_sheet_name(base[: 31 - len(suffix)] + suffix)
            n += 1
        used.add(name.lower())
        sheet_names.append(name)

    if "점검대상" in wb.sheetnames:
        ws = wb["점검대상"]
        for i, (report, sheet_name) in enumerate(zip(reports, sheet_names)):
            row = 7 + i
            ws.cell(row, 2).value = i + 1
            ws.cell(row, 3).value = sheet_name
            ws.cell(row, 4).value = report.target_os
            ws.cell(row, 5).value = report.ip_address
            ws.cell(row, 7).value = 3
            ws.cell(row, 8).value = 3
            ws.cell(row, 9).value = 3
            ws.cell(row, 10).value = (
                f'=IF(AND(SUM(G{row}:I{row})>=8, SUM(G{row}:I{row})<=9), "1등급", '
                f'IF(AND(SUM(G{row}:I{row})>=6, SUM(G{row}:I{row})<=7), "2등급", '
                f'IF(AND(SUM(G{row}:I{row})>=3, SUM(G{row}:I{row})<=5), "3등급", "")))'
            )

    sample_name = "sample" if "sample" in wb.sheetnames else None
    if sample_name:
        for report, sheet_name in zip(reports, sheet_names):
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.copy_worksheet(wb[sample_name])
            ws.title = sheet_name
            for coord, value in (("D4", report.hostname), ("D5", report.ip_address), ("H4", report.target_os)):
                try:
                    ws[coord] = value
                except Exception:
                    pass
            _fill_host_detail_sheet(
                ws,
                report,
                code_col=4,
                result_col=13,
                result_mapper=_status_yn,
                evidence_col=15,
                op_col=None,
                start_row=21,
                end_row=87,
            )
        try:
            del wb[sample_name]
        except Exception:
            pass

    wb.save(out)
    wb.close()
    return out


def export_reports(
    reports: list[HostReport],
    options: ExportOptions,
    guidelines: list[dict] | None = None,
) -> ExportResult:
    logs: list[str] = []
    files: list[str] = []
    now = datetime.now().strftime("%H:%M:%S")
    logs.append(f"[{now}] 보고서 생성 작업을 시작합니다...")

    if not (options.detail or options.summary or options.action_plan):
        raise ValueError("출력할 보고서 양식을 최소 하나 이상 선택해야 합니다.")
    if not (options.unix or options.dbms or options.win_server or options.pc):
        raise ValueError("내보낼 자산 유형을 최소 하나 이상 선택해야 합니다.")

    target = _filter_reports(reports, options)
    if not target:
        raise ValueError("선택한 자산 유형에 해당하는 진단 결과 호스트가 존재하지 않습니다.")

    guides = guidelines or []

    # Safety net if caller did not persist enrichment yet.
    try:
        from app.config import UPLOAD_DIR
        from app.services.json_parser import enrich_reports_from_uploads

        n = enrich_reports_from_uploads(target, UPLOAD_DIR, guides)
        if n:
            logs.append(f"[{datetime.now():%H:%M:%S}] 원본 JSON에서 점검결과/설정값을 복구했습니다. ({n} items)")
    except Exception:
        pass

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    groups = _group_by_os(target)

    if options.detail:
        for os_type, group in groups.items():
            logs.append(f"[{datetime.now():%H:%M:%S}] {os_type} 상세결과보고서 생성 중... ({len(group)} hosts)")
            path = generate_detailed_report(group, os_type, guidelines=guides)
            files.append(path.name)
            logs.append(f"[{datetime.now():%H:%M:%S}]   - 완료: {path.name}")

    if options.summary:
        for os_type, group in groups.items():
            logs.append(f"[{datetime.now():%H:%M:%S}] {os_type} 위험관리 계획서 생성 중... ({len(group)} hosts)")
            path = generate_risk_workbook(
                group,
                os_type,
                "UNIX_서버_위험관리 계획서_양식.xlsm",
                "서버_위험관리_계획서",
            )
            files.append(path.name)
            logs.append(f"[{datetime.now():%H:%M:%S}]   - 완료: {path.name}")

    if options.action_plan:
        for os_type, group in groups.items():
            logs.append(f"[{datetime.now():%H:%M:%S}] {os_type} 위험 분석 평가표 생성 중... ({len(group)} hosts)")
            path = generate_risk_workbook(
                group,
                os_type,
                "UNIX_서버_위험_분석_평가표_양식.xlsm",
                "서버_위험_분석_평가표",
            )
            files.append(path.name)
            logs.append(f"[{datetime.now():%H:%M:%S}]   - 완료: {path.name}")

    logs.append(f"[{datetime.now():%H:%M:%S}] 모든 보고서 생성이 완료되었습니다.")
    return ExportResult(files=files, logs=logs)
