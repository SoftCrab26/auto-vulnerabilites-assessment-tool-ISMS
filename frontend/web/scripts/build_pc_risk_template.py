#!/usr/bin/env python3
"""Build PC 위험분석 및 평가 보고서 template from the UNIX risk-plan workbook."""

from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "templates_excel" / "UNIX_서버_위험관리 계획서_양식.xlsm"
DST = ROOT / "templates_excel" / "PC_위험분석및평가보고서_양식.xlsm"
GUIDE = ROOT.parent.parent / "guide" / "주통기" / "주통기_pc.md"

ITEMS: list[tuple[str, str, str, str, str]] = []
# code, title, severity, pass_criteria, fail_criteria

RISKS = [
    (
        "비밀번호가 장기간 사용되어 유출·재사용될 위험이 존재",
        "최대 암호 사용 기간을 90일 이하로 설정",
    ),
    (
        "짧은 암호 또는 복잡성 미적용으로 무차별 대입 공격에 취약",
        "최소 암호 길이 8자 이상 및 암호 복잡성 설정",
    ),
    (
        "복구 콘솔 자동 로그온으로 비인가자가 시스템에 접근할 수 있음",
        "복구 콘솔 자동 로그온을 사용 안 함으로 설정",
    ),
    (
        "기본 공유(C$, ADMIN$ 등)를 통한 원격 접근 위험이 존재",
        "기본 공유 및 불필요한 공유 폴더 제거",
    ),
    (
        "불필요한 서비스가 자동 시작되어 공격면이 증가",
        "불필요한 서비스를 사용 안 함으로 설정",
    ),
    (
        "비인가 상용 메신저를 통한 정보 유출 위험이 존재",
        "비인가 상용 메신저 설치 금지 및 제거",
    ),
    (
        "NTFS가 아닌 파일시스템은 접근권한 통제가 어려움",
        "고정 드라이브를 NTFS로 설정",
    ),
    (
        "멀티 부팅 환경에서 보안 정책을 우회할 수 있음",
        "PC에 OS를 하나만 설치",
    ),
    (
        "임시 인터넷 파일에 민감정보가 잔존할 수 있음",
        "브라우저 종료 시 임시 인터넷 파일 삭제 설정",
    ),
    (
        "보안 패치 미적용으로 알려진 취약점에 노출",
        "HOT FIX 설치 및 자동 업데이트 유지",
    ),
    (
        "지원이 종료된 OS는 보안 업데이트를 받을 수 없음",
        "지원되는 Windows OS Build 적용",
    ),
    (
        "자동 로그온 설정 시 계정 정보가 노출될 수 있음",
        "AutoAdminLogon 비활성화",
    ),
    (
        "백신 미설치 시 악성코드 감염 위험이 증가",
        "바이러스 백신 프로그램 설치 및 주기적 업데이트",
    ),
    (
        "실시간 감시 미사용 시 악성코드 탐지가 지연",
        "백신 실시간 감시 기능 활성화",
    ),
    (
        "방화벽 미사용 시 외부 침입 위험이 증가",
        "Windows 방화벽을 사용으로 설정",
    ),
    (
        "화면 잠금 미설정 시 자리 비움 중 무단 사용 가능",
        "화면보호기 암호 보호 및 대기 시간 10분 이하 설정",
    ),
    (
        "이동식 미디어 자동 실행으로 악성코드가 유입될 수 있음",
        "Autorun 방지 정책 설정",
    ),
    (
        "원격 지원 허용 시 외부에서 시스템을 장악할 위험이 존재",
        "원격 지원을 사용 안 함으로 설정",
    ),
]

TT_CODES = [
    "TT-07",
    "TT-07",
    "TT-14",
    "TT-14",
    "TT-15",
    "TT-15",
    "TT-13",
    "TT-13",
    "TT-13",
    "TT-15",
    "TT-15",
    "TT-14",
    "TT-15",
    "TT-15",
    "TT-15",
    "TT-14",
    "TT-14",
    "TT-14",
]

CAT_STARTS = {
    0: "계정 관리",
    3: "접근 관리",
    9: "패치 관리",
    11: "보안 관리",
}


def _load_items() -> None:
    text = GUIDE.read_text(encoding="utf-8")
    for line in text.splitlines():
        if not line.strip().startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) < 5:
            continue
        m = re.search(r"PC-(\d+)", cells[0])
        if not m:
            continue
        title = cells[1]
        sev = cells[2]
        good = cells[3]
        bad = cells[4]
        ITEMS.append((f"PC-{int(m.group(1)):02d}", title, sev, good, bad))
    if len(ITEMS) != 18:
        raise SystemExit(f"expected 18 PC items, got {len(ITEMS)}")


def _unmerge_all(ws, coords: list[str]) -> None:
    for coord in coords:
        try:
            ws.unmerge_cells(coord)
        except Exception:
            pass


def _replace_cf(ws, old: str, new: str) -> None:
    rules = list(ws.conditional_formatting._cf_rules.items())
    keep = []
    for key, rule_list in rules:
        ref = str(getattr(key, "sqref", key) or "")
        if ref == old:
            try:
                del ws.conditional_formatting._cf_rules[key]
            except Exception:
                pass
            for rule in rule_list:
                ws.conditional_formatting.add(new, rule)
        else:
            keep.append((key, rule_list))


def _fix_validations(ws, new_end: int) -> None:
    if not ws.data_validations:
        return
    for dv in ws.data_validations.dataValidation:
        formula = str(dv.formula1 or "")
        if "Y,N" in formula.replace('"', ""):
            dv.sqref = f"M21:M{new_end}"
        elif "1,2,3" in formula.replace('"', ""):
            dv.sqref = f"Z21:AB{new_end}"
        elif "HOT FIX" in formula.replace('"', ""):
            dv.sqref = f"AD21:AF{new_end} AG21:AG33"


def _status_text(good: str) -> str:
    body = good.rstrip(".")
    if "경우" not in body:
        body = f'{body} 경우'
    return f'■ 기준 : {body} "양호"\n\n■ 현황 : '


def build() -> None:
    _load_items()
    shutil.copy2(SRC, DST)
    wb = load_workbook(DST, keep_vba=True)

    if "표지" in wb.sheetnames:
        wb["표지"]["B12"] = "PC 위험분석 및 평가 보고서"

    ws = wb["점검대상"]
    if ws["B2"].value:
        ws["B2"] = "◎ PC 점검대상"
    ws["D7"] = "Windows"
    ws["F7"] = "개인 PC"

    # --- 요약 통계 ---
    sm = wb["요약 통계"]
    _unmerge_all(sm, ["B6:B18", "B19:B38", "B39:B68", "B70:B72"])
    for i, (code, title, sev, _g, _b) in enumerate(ITEMS):
        row = 6 + i
        sm.cell(row, 3).value = code
        sm.cell(row, 4).value = title
        sm.cell(row, 5).value = sev
        if i in CAT_STARTS:
            sm.cell(row, 2).value = CAT_STARTS[i]
        elif i not in (0,):
            # keep first-of-group only
            if i not in CAT_STARTS:
                sm.cell(row, 2).value = None
    for start, end in ((6, 8), (9, 14), (15, 16), (17, 23)):
        try:
            sm.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
        except Exception:
            pass
    sm["B6"] = "계정 관리"
    sm["B9"] = "접근 관리"
    sm["B15"] = "패치 관리"
    sm["B17"] = "보안 관리"

    for row in range(24, 73):
        for col in range(2, 13):
            sm.cell(row, col).value = None

    for col, letter in enumerate(["F", "G", "H", "I", "J", "K"], start=6):
        sm.cell(73, col).value = f"=SUM({letter}6:{letter}23)"

    sm.cell(73, 12).value = ArrayFormula("L73", '=SUM(IF(L$6:L$23<>"N/A",$F$6:$F$23))')
    sm.cell(74, 12).value = ArrayFormula("L74", '=SUMIF(L$6:L$23,"=N",$F$6:$F$23)')
    for row in (75, 76, 77):
        sm.cell(row, 12).value = f"=COUNTIF(L$6:L$23,$J{row})"

    # --- 보안수준 통계 ---
    sec = wb["보안수준 통계"]
    ranges = [
        (6, "B6", "G6:G8", "H6:H8", "K6:K8", "I6:J8", "J6:J8"),
        (7, "B9", "G9:G14", "H9:H14", "K9:K14", "I9:J14", "J9:J14"),
        (8, "B15", "G15:G16", "H15:H16", "K15:K16", "I15:J16", "J15:J16"),
        (9, "B17", "G17:G23", "H17:H23", "K17:K23", "I17:J23", "J17:J23"),
    ]
    for row, bref, g, h, k, ij, j in ranges:
        sec.cell(row, 2).value = f"='요약 통계'!{bref}"
        sec.cell(row, 4).value = f"=SUM('요약 통계'!{g})"
        sec.cell(row, 5).value = f"=SUM('요약 통계'!{h})"
        sec.cell(row, 6).value = f"=SUM('요약 통계'!{k})"
        sec.cell(row, 7).value = f"=SUM('요약 통계'!{ij})"
        sec.cell(row, 8).value = f"=SUM('요약 통계'!{j})"
    sec.cell(10, 2).value = "-"
    for col in range(4, 9):
        sec.cell(10, col).value = 0
    sec.cell(10, 9).value = '="N/A"'

    # --- 잠재위험 DB ---
    db = wb["잠재위험 및 대응책 DB"]
    db["A1"] = "잠재위험"
    db["B1"] = "대응책"
    for i, (risk, action) in enumerate(RISKS):
        db.cell(2 + i, 1).value = risk
        db.cell(2 + i, 2).value = action
    for row in range(20, (db.max_row or 20) + 1):
        db.cell(row, 1).value = None
        db.cell(row, 2).value = None

    def _fill_host_sheet(ws) -> None:
        _unmerge_all(
            ws,
            ["B21:C33", "B34:C53", "B54:C83", "B84:C84", "B85:C87"],
        )
        for i, (code, title, sev, good, _bad) in enumerate(ITEMS):
            row = 21 + i
            sum_row = 6 + i
            db_row = 2 + i
            if i in CAT_STARTS:
                ws.cell(row, 2).value = f"='요약 통계'!B{sum_row}"
            else:
                ws.cell(row, 2).value = None
            ws.cell(row, 4).value = code
            ws.cell(row, 5).value = f"='요약 통계'!D{sum_row}"
            ws.cell(row, 10).value = f"='요약 통계'!E{sum_row}"
            ws.cell(row, 11).value = f'=IF(M{row}="N/A","-",\'요약 통계\'!F{sum_row})'
            ws.cell(row, 12).value = TT_CODES[i]
            ws.cell(row, 15).value = _status_text(good)
            ws.cell(row, 16).value = f'=IF(M{row}="N",\'잠재위험 및 대응책 DB\'!A{db_row},"-")'
            ws.cell(row, 17).value = f'=IF(M{row}="N",\'잠재위험 및 대응책 DB\'!B{db_row},"-")'
        try:
            ws.merge_cells("B21:C23")
            ws.merge_cells("B24:C29")
            ws.merge_cells("B30:C31")
            ws.merge_cells("B32:C38")
        except Exception:
            pass
        leftover = [str(m) for m in ws.merged_cells.ranges if m.max_row >= 39]
        _unmerge_all(ws, leftover)
        if ws.max_row and ws.max_row >= 39:
            ws.delete_rows(39, 49)
        _replace_cf(ws, "D21:D87", "D21:D38")
        _replace_cf(ws, "W21:W87", "W21:W38")
        _replace_cf(ws, "M21:M87", "M21:M38")
        _fix_validations(ws, 38)

    _fill_host_sheet(wb["sample"])
    if "Origin(숨김처리)" in wb.sheetnames:
        _fill_host_sheet(wb["Origin(숨김처리)"])
    if "잠재위험 및 대응책 DB" in wb.sheetnames:
        wb["잠재위험 및 대응책 DB"].sheet_state = "visible"

    wb.save(DST)
    wb.close()
    print("wrote", DST)


if __name__ == "__main__":
    build()
