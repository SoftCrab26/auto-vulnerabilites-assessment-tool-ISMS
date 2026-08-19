#!/usr/bin/env python3
"""Build 정보보호시스템 위험분석 및 평가 보고서 template from the UNIX workbook."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "templates_excel" / "UNIX_서버_위험관리 계획서_양식.xlsm"
DST = ROOT / "templates_excel" / "정보보호시스템_위험분석및평가보고서_양식.xlsm"

# code, title, severity, 양호 기준
ITEMS: list[tuple[str, str, str, str]] = [
    ("S-01", "보안장비 Default 계정 변경", "상", "장비에서 제공하고 있는 디폴트 계정 명을 변경하여 사용하는 경우 (ID 변경이 불가능할 경우 패스워드로 보완필요)"),
    ("S-02", "보안장비 Default 패스워드 변경", "상", "비밀번호 관리 정책에 맞는 비밀번호가 사용된 경우"),
    ("S-03", "보안장비 계정별 권한 설정", "상", "사용자 별 계정의 용도파악 및 적절한 권한을 부여하는 경우"),
    ("S-04", "보안장비 계정 관리", "상", "불필요한 공용계정 및 휴면계정을 제거하거나 관리하는 경우"),
    ("S-05", "로그인 실패횟수 제한", "상", "로그온 실패 수를 5회 이하로 제한한 경우"),
    ("S-06", "보안장비 원격 관리 접근 통제", "상", "원격관리 시 관리자 IP만 접근 가능하도록 설정한 경우"),
    ("S-07", "보안장비 보안 접속", "상", "보안장비 접속 시 암호화 통신을 하는 경우"),
    ("S-08", "Session timeout 설정", "상", "Session Timeout 시간을 설정한 경우"),
    ("S-09", "벤더에서 제공하는 최신 업데이트 적용", "상", "패치적용정책을 수립하여 주기적으로 패치를 관리하고 있을 경우"),
    ("S-10", "보안장비 로그 설정", "중", "정책에 따른 로그설정이 되어있는 경우"),
    ("S-11", "보안장비 로그 보관", "중", "정책에 따라 로그보관 설정이 되어있는 경우"),
    ("S-12", "보안장비 정책 백업 설정", "중", "보안장비에 적용된 정책을 별도의 파일로 보관하고 있는 경우"),
    ("S-13", "원격 로그 서버 사용", "중", "별도의 로그서버를 구축하여 통합 관리하는 경우"),
    ("S-14", "NTP 서버 연동", "중", "NTP 및 시간 동기화 설정이 되어 있는 경우"),
    ("S-15", "정책 관리", "상", "정책에 대한 주기적인 검사로 미 사용 및 중복된 정책을 확인하여 제거하는 경우"),
    ("S-16", "NAT 설정", "상", "NAT 설정으로 내부네트워크를 보호하는 경우"),
    ("S-17", "DMZ 설정", "상", "DMZ를 구성하여 내부네트워크를 보호하는 경우"),
    ("S-18", "최소한의 서비스만 제공", "상", "all deny 설정을 하고, 방화벽에 최소서비스만 허용할 경우"),
    ("S-19", "이상징후 탐지 경고 기능 설정", "상", "이상징후 탐지 시 관리자에게 이메일이나 SMS로 통보되는 경우"),
    ("S-20", "장비 사용량 검토", "상", "장비사용량을 정기적으로 모니터링 및 검토할 경우"),
    ("S-21", "SNMP 서비스 확인", "상", "SNMP 서비스를 불필요하게 사용하지 않는 경우"),
    ("S-22", "SNMP Community String 복잡성 설정", "상", "SNMP 서비스를 사용하지 않거나, 유추하기 어려운 community string을 설정한 경우"),
    ("S-23", "유해 트래픽 차단 정책 설정", "중", "유해트래픽 차단정책이 설정되어 있는 경우"),
]

RISKS = [
    ("디폴트 계정을 통한 비인가 접근 위험이 존재", "디폴트 계정명을 변경하거나 패스워드로 보완"),
    ("취약한 비밀번호로 계정 탈취 위험이 존재", "비밀번호 관리 정책에 맞는 비밀번호 사용"),
    ("과도한 권한 부여로 오남용 위험이 존재", "계정 용도에 맞는 최소 권한 부여"),
    ("공용·휴면 계정으로 책임추적성이 저하됨", "불필요 공용계정 및 휴면계정 제거"),
    ("로그인 실패 제한 미흡 시 무차별 대입 공격에 노출", "로그온 실패 횟수를 5회 이하로 제한"),
    ("원격관리 접근 통제 미흡 시 외부 침입 위험이 존재", "원격관리를 관리자 IP로 제한"),
    ("비암호화 접속 시 인증정보가 노출될 수 있음", "SSH/HTTPS 등 암호화 통신 사용"),
    ("세션 타임아웃 미설정 시 방치 세션이 악용될 수 있음", "Session Timeout 시간 설정"),
    ("보안 패치 미적용으로 알려진 취약점에 노출", "벤더 패치 정책을 수립하고 주기적으로 적용"),
    ("로그 미설정 시 침해사고 추적이 어려움", "정책에 따른 로그 설정"),
    ("로그 보관 미흡 시 증거 데이터가 소실됨", "정책에 따른 로그 보관 설정"),
    ("정책 백업 부재 시 장애·침해 후 복구가 어려움", "적용 정책을 별도 파일로 백업"),
    ("로컬 로그만 사용 시 로그 위변조 위험이 존재", "원격 로그 서버로 통합 관리"),
    ("시간 동기화 미설정 시 로그 상관분석이 어려움", "NTP 서버 연동 및 시각 동기화"),
    ("미사용·중복 정책으로 우회 경로가 남을 수 있음", "정책을 주기적으로 점검하고 정리"),
    ("NAT 미구성 시 내부망 주소가 노출될 수 있음", "NAT로 내부 네트워크 보호"),
    ("DMZ 미구성 시 외부 서비스가 내부망과 혼재됨", "DMZ를 구성하여 내부망 분리"),
    ("과도한 허용 정책으로 공격면이 증가", "all deny 후 최소 서비스만 허용"),
    ("이상징후 미통보 시 침해 대응이 지연됨", "이상징후 탐지 시 이메일/SMS 통보"),
    ("사용량 미검토 시 장애·공격 징후를 놓칠 수 있음", "CPU/메모리/트래픽 등 사용량을 정기 검토"),
    ("불필요 SNMP로 장비 정보가 노출될 수 있음", "불필요한 SNMP 서비스 비활성화"),
    ("기본 Community String 사용 시 SNMP가 악용됨", "유추가 어려운 Community String 설정"),
    ("유해 트래픽 차단 미흡 시 내부망이 공격에 노출", "유해 트래픽 차단 정책 설정"),
]

TT_CODES = (
    ["TT-07"] * 5
    + ["TT-14"] * 3
    + ["TT-15"]
    + ["TT-15"] * 5
    + ["TT-14"] * 9
)

CAT_STARTS = {
    0: "계정 관리",
    5: "접근 관리",
    8: "패치 관리",
    9: "로그 관리",
    14: "기능 관리",
}
CAT_MERGES = ((21, 25), (26, 28), (29, 29), (30, 34), (35, 43))
ITEM_END = 21 + len(ITEMS) - 1  # 43
SUM_END = 6 + len(ITEMS) - 1  # 28


def _unmerge_all(ws, coords: list[str]) -> None:
    for coord in coords:
        try:
            ws.unmerge_cells(coord)
        except Exception:
            pass


def _replace_cf(ws, old: str, new: str) -> None:
    for key, rule_list in list(ws.conditional_formatting._cf_rules.items()):
        ref = str(getattr(key, "sqref", key) or "")
        if ref != old:
            continue
        try:
            del ws.conditional_formatting._cf_rules[key]
        except Exception:
            pass
        for rule in rule_list:
            ws.conditional_formatting.add(new, rule)


def _fix_validations(ws, new_end: int) -> None:
    if not ws.data_validations:
        return
    for dv in ws.data_validations.dataValidation:
        formula = str(dv.formula1 or "")
        if "Y,N" in formula.replace('"', ""):
            dv.sqref = f"M21:M{new_end}"
        elif "1,2,3" in formula.replace('"', ""):
            dv.sqref = f"Z21:AB{new_end}"
        elif "HOT FIX" in formula.replace('"', ""):
            dv.sqref = f"AD21:AF{new_end} AG21:AG33"


def _status_text(good: str) -> str:
    body = good.rstrip(".")
    if "경우" not in body:
        body = f"{body} 경우"
    return f'■ 기준 : {body} "양호"\n\n■ 현황 : '


def build() -> None:
    shutil.copy2(SRC, DST)
    wb = load_workbook(DST, keep_vba=True)

    if "표지" in wb.sheetnames:
        wb["표지"]["B12"] = "정보보호시스템 위험분석 및 평가 보고서"

    ws = wb["점검대상"]
    if ws["B2"].value:
        ws["B2"] = "◎ 정보보호시스템 점검대상"
    ws["D7"] = "방화벽"
    ws["F7"] = "보안장비"

    sm = wb["요약 통계"]
    _unmerge_all(sm, ["B6:B18", "B19:B38", "B39:B68", "B70:B72"])
    for i, (code, title, sev, _g) in enumerate(ITEMS):
        row = 6 + i
        sm.cell(row, 3).value = code
        sm.cell(row, 4).value = title
        sm.cell(row, 5).value = sev
        sm.cell(row, 2).value = CAT_STARTS.get(i)
    for start, end in ((6, 10), (11, 13), (14, 14), (15, 19), (20, 28)):
        try:
            sm.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
        except Exception:
            pass
    sm["B6"] = "계정 관리"
    sm["B11"] = "접근 관리"
    sm["B14"] = "패치 관리"
    sm["B15"] = "로그 관리"
    sm["B20"] = "기능 관리"

    for row in range(SUM_END + 1, 73):
        for col in range(2, 13):
            sm.cell(row, col).value = None

    for col, letter in enumerate(["F", "G", "H", "I", "J", "K"], start=6):
        sm.cell(73, col).value = f"=SUM({letter}6:{letter}{SUM_END})"

    sm.cell(73, 12).value = ArrayFormula("L73", f'=SUM(IF(L$6:L${SUM_END}<>"N/A",$F$6:$F${SUM_END}))')
    sm.cell(74, 12).value = ArrayFormula("L74", f'=SUMIF(L$6:L${SUM_END},"=N",$F$6:$F${SUM_END})')
    for row in (75, 76, 77):
        sm.cell(row, 12).value = f"=COUNTIF(L$6:L${SUM_END},$J{row})"

    sec = wb["보안수준 통계"]
    ranges = [
        (6, "B6", "G6:G10", "H6:H10", "K6:K10", "I6:J10", "J6:J10"),
        (7, "B11", "G11:G13", "H11:H13", "K11:K13", "I11:J13", "J11:J13"),
        (8, "B14", "G14:G14", "H14:H14", "K14:K14", "I14:J14", "J14:J14"),
        (9, "B15", "G15:G19", "H15:H19", "K15:K19", "I15:J19", "J15:J19"),
        (10, "B20", "G20:G28", "H20:H28", "K20:K28", "I20:J28", "J20:J28"),
    ]
    for row, bref, g, h, k, ij, j in ranges:
        sec.cell(row, 2).value = f"='요약 통계'!{bref}"
        sec.cell(row, 4).value = f"=SUM('요약 통계'!{g})"
        sec.cell(row, 5).value = f"=SUM('요약 통계'!{h})"
        sec.cell(row, 6).value = f"=SUM('요약 통계'!{k})"
        sec.cell(row, 7).value = f"=SUM('요약 통계'!{ij})"
        sec.cell(row, 8).value = f"=SUM('요약 통계'!{j})"

    db = wb["잠재위험 및 대응책 DB"]
    db["A1"] = "잠재위험"
    db["B1"] = "대응책"
    for i, (risk, action) in enumerate(RISKS):
        db.cell(2 + i, 1).value = risk
        db.cell(2 + i, 2).value = action
    for row in range(2 + len(RISKS), (db.max_row or 20) + 1):
        db.cell(row, 1).value = None
        db.cell(row, 2).value = None

    def _fill_host_sheet(ws) -> None:
        _unmerge_all(
            ws,
            ["B21:C33", "B34:C53", "B54:C83", "B84:C84", "B85:C87"],
        )
        for i, (code, title, sev, good) in enumerate(ITEMS):
            row = 21 + i
            sum_row = 6 + i
            db_row = 2 + i
            ws.cell(row, 2).value = f"='요약 통계'!B{sum_row}" if i in CAT_STARTS else None
            ws.cell(row, 4).value = code
            ws.cell(row, 5).value = f"='요약 통계'!D{sum_row}"
            ws.cell(row, 10).value = f"='요약 통계'!E{sum_row}"
            ws.cell(row, 11).value = f'=IF(M{row}="N/A","-",\'요약 통계\'!F{sum_row})'
            ws.cell(row, 12).value = TT_CODES[i]
            ws.cell(row, 15).value = _status_text(good)
            ws.cell(row, 16).value = f'=IF(M{row}="N",\'잠재위험 및 대응책 DB\'!A{db_row},"-")'
            ws.cell(row, 17).value = f'=IF(M{row}="N",\'잠재위험 및 대응책 DB\'!B{db_row},"-")'
        for start, end in CAT_MERGES:
            try:
                ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=3)
            except Exception:
                pass
        leftover = [str(m) for m in ws.merged_cells.ranges if m.max_row >= ITEM_END + 1]
        _unmerge_all(ws, leftover)
        if ws.max_row and ws.max_row >= ITEM_END + 1:
            ws.delete_rows(ITEM_END + 1, 50)
        _replace_cf(ws, "D21:D87", f"D21:D{ITEM_END}")
        _replace_cf(ws, "W21:W87", f"W21:W{ITEM_END}")
        _replace_cf(ws, "M21:M87", f"M21:M{ITEM_END}")
        _fix_validations(ws, ITEM_END)

    _fill_host_sheet(wb["sample"])
    if "Origin(숨김처리)" in wb.sheetnames:
        _fill_host_sheet(wb["Origin(숨김처리)"])
    if "잠재위험 및 대응책 DB" in wb.sheetnames:
        wb["잠재위험 및 대응책 DB"].sheet_state = "visible"

    wb.save(DST)
    wb.close()
    print("wrote", DST)


if __name__ == "__main__":
    build()
